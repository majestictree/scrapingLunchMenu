import copy
import time
import requests, bs4
import openpyxl
from openpyxl.drawing.image import Image
from pathlib import Path

#画像フォルダを用意
output_folder = Path('画像')

#メニューとなるエクセルブック作成
wb = openpyxl.load_workbook('scraping_excel.xlsx')
sheet = wb.active
sheet.title = '甚兵衛'

#関数定義
def scrapingdataToList(scrapingdata, op):
    rtn_list = []
    for data in scrapingdata:
        if (op == 'text'):
            rtn_list.append(data.text)
        elif (op == 'src'):
            rtn_list.append(data.get('src'))
    return rtn_list

def downloadImages(url_list):
    output_folder.mkdir(exist_ok=True)
    file_i = 1
    for item in images_url_list:
        url = "https:" + item
        filename = str(file_i) + ".jpg"
        print(url)
        save_path = output_folder.joinpath(filename)
        time.sleep(1.0)
        try:
            image = requests.get(url)
            open(save_path, 'wb').write(image.content)
            print(save_path)
        except ValueError:
            print("ValueError!")
        file_i += 1

#スクレイピング処理開始
res = requests.get('https://demae-can.com/shop/menu/3005065/')
res.raise_for_status()
soup = bs4.BeautifulSoup(res.text, 'html.parser')
names = soup.select('.item_name')
prices = soup.select('.item_price')
images_url = soup.select('.item_img img')

name_list = scrapingdataToList(names, 'text')
price_list = scrapingdataToList(prices, 'text')
images_url_list = scrapingdataToList(images_url, 'src')

#値段が550円よりも高いメニューを除外
price = 0
price_list_tmp = copy.copy(price_list)
for i in reversed(range(0, 80)):
    price = int(price_list_tmp[i].strip(chr(165)).replace(',',''))
    if (price > 550):
        name_list.pop(i)
        price_list.pop(i)
        price_list_tmp.pop(i)
        images_url_list.pop(i)

#画像をローカルに保存する
downloadImages(images_url_list)

#45個目までのメニューをエクセルに記述する
#メニュー名と値段
flag = 0
j = 2
k = 2
for i in range(1,45):
    sheet.cell(row = j, column = k, value = name_list[i-1])
    sheet.cell(row = j+1, column = k, value = price_list[i-1])
    if (flag == 0):
        flag = 1
        k = 4
    else:
        flag = 0
        j += 2
        k = 2

#画像の貼り付け
flag = 0
j = 2
k = 1
for i in range(1,45):
    img = Image('C:\\Users\\user\\Desktop\\WorkSpace\\1.ToDo\\scrapingLunchMenu\\画像\\' + str(i) + ".jpg")
    cell_address = sheet.cell(row = j, column = k).coordinate
    img.anchor = cell_address
    sheet.add_image(img)
    if (flag == 0):
        flag = 1
        k = 3
    else:
        flag = 0
        j += 2
        k = 1

wb.save('scraping_excel.xlsx')
wb.close